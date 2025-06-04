from openpyxl import load_workbook
import json
import requests
import time
import re
import random
from bs4 import BeautifulSoup
import datetime
import sys

# 创建日志记录函数，替代print
original_print = print
log_file = None

def log_print(*args, **kwargs):
    # 获取当前时间
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # 构建带时间戳的消息
    message = f"[{timestamp}] " + " ".join(map(str, args))
    # 调用原始print输出到控制台
    original_print(message, **kwargs)
    # 写入日志文件
    if log_file:
        log_file.write(message + "\n")
        log_file.flush()  # 确保立即写入文件

# 替换全局print函数
print = log_print

def get_stock_code(company_name):
    """通过新浪API查询股票代码"""
    try:
        url = f"http://suggest3.sinajs.cn/suggest/type=11&key={company_name}"
        response = requests.get(url)
        if response.status_code == 200:
            data = response.text.split('="')[1].split(';')[0]
            if data:
                parts = data.split(',')
                if len(parts) > 3:
                    return parts[3]  # 返回股票代码
        time.sleep(0.5)
    except Exception as e:
        print(f"查询{company_name}股票代码失败: {e}")
    return None

def set_report_link(_title, reports, file_path):
    """根据返回数据找到符合要求的PDF链接"""
    title = _title.replace(' ', '')
    if '年度报告' in title and '摘要' not in title and '英文' not in title:
        year_match = re.search(r'(\d{4})年', title)
        if year_match:
            year = year_match.group(1)
            if year in reports:
                print(f"  title: {title}, pdf_url: {file_path}")
                reports[year] = file_path

def get_report_links(stock_code, company_name, current_index, total_count):
    """根据股票代码获取年报PDF链接"""
    if not stock_code:
        return {}
        
    reports = {str(year): "" for year in range(2018, 2025)}
    max_retries = 2
    timeout = 10
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'application/json'
    }
    
    try:
        for _ in range(max_retries):
            try:
                if stock_code.startswith('sh'):
                    # 上交所年报查询 - 使用API接口
                    stock_num = stock_code[2:]
                    api_url = "https://query.sse.com.cn/security/stock/queryCompanyBulletin.do"
                    params = {
                        "jsonCallBack": "jsonpCallback83303800",
                        "isPagination": "true",
                        "pageHelp.pageSize": 50,
                        "pageHelp.pageNo": 1,
                        "pageHelp.beginPage": 1,
                        "pageHelp.cacheSize": 1,
                        "pageHelp.endPage": 1,
                        "productId": stock_num,
                        "securityType": "0101,120100,020100,020200,120200",
                        "reportType2": "DQBG",
                        "reportType": "YEARLY",
                        "beginDate": "",
                        "endDate": "",
                        "_": str(int(time.time() * 1000))
                    }
                    api_headers = headers.copy()
                    api_headers['Referer'] = 'https://www.sse.com.cn'
                    res = requests.get(api_url, params=params, headers=api_headers, timeout=timeout)
                    res.raise_for_status()
                    # "\033[32m这是绿色字体\033[0m"
                    print(f"\033[32m[{current_index}/{total_count}] {company_name}({stock_code}) 上交所API请求状态: 200, 返回数据: {res.text[:40]}...\033[0m")
                    
                    # 解析JSONP响应
                    json_str = res.text[len(params['jsonCallBack'])+1:-1]
                    data = json.loads(json_str)
                    result = data.get('result', [])
                    if len(result) > 0:
                        for item in result:
                            title = item.get('TITLE', '')
                            set_report_link(title, reports, f"http://www.sse.com.cn{item['URL']}")
                            
                elif stock_code.startswith('sz'):
                    # 深交所年报查询 - 使用API接口
                    stock_num = stock_code[2:]
                    api_url = "https://www.szse.cn/api/disc/announcement/annList"
                    params = {
                        "random": random.random(),
                        "seDate": ["", ""],
                        "stock": [stock_num],
                        "channelCode": ["fixed_disc"],
                        "pageSize": 50,
                        "pageNum": 1
                    }
                    res = requests.post(api_url, json=params, headers=headers, timeout=timeout)
                    res.raise_for_status()
                    print(f"\033[32m[{current_index}/{total_count}] {company_name}({stock_code}) 深交所API请求状态: 200, 返回数据: {res.text[:40]}...\033[0m")
                    data = res.json()
                    dataList = data.get('data', [])
                    if len(dataList) > 0:
                        for item in dataList:
                            title = item.get('title', '')
                            set_report_link(title, reports, f"https://disc.static.szse.cn/download{item['attachPath']}")
                
                time.sleep(3)  # 控制请求频率
                break  # 成功则跳出重试循环
                
            except requests.exceptions.RequestException as e:
                print(f"[{current_index}/{total_count}] 获取{company_name}({stock_code})年报链接失败(重试中): {e}")
                print(f"[{current_index}/{total_count}] {company_name}({stock_code}) {'上交所' if stock_code.startswith('sh') else '深交所'}API请求状态: 失败")
                time.sleep(5)
                continue
                
    except Exception as e:
        print(f"[{current_index}/{total_count}] 获取{company_name}({stock_code})年报链接最终失败: {e}")
    
    return reports

def read_excel_to_dict(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    
    result = {}
    sector_states = [{'main': None, 'sub': None} for _ in range(9)]
    
    # 先计算总查询数量
    total_companies = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        for group_idx in range(9):
            col_start = group_idx * 3
            if col_start + 2 >= len(row):
                continue
            company = row[col_start+2]
            if company and str(company).strip():
                total_companies += 1
    
    # 重置文件指针
    ws = wb.active
    current_index = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        for group_idx in range(9):
            col_start = group_idx * 3
            if col_start + 2 >= len(row):
                continue
                
            main_sector = row[col_start]
            sub_sector = row[col_start+1]
            company = row[col_start+2]
            
            state = sector_states[group_idx]
            if main_sector is not None:
                state['main'] = str(main_sector).strip()
            if sub_sector is not None:
                state['sub'] = str(sub_sector).strip()
                
            if None in (state['main'], state['sub'], company):
                continue
                
            company = str(company).strip()
            if not company:
                continue
                
            # 查询股票代码
            stock_code = get_stock_code(company)
            
            current_index += 1
            company_data = {
                "name": company,
                "code": stock_code or "",
                "reports": get_report_links(stock_code, company, current_index, total_companies) if stock_code else {}
            }
                
            if state['main'] not in result:
                result[state['main']] = {}
            if state['sub'] not in result[state['main']]:
                result[state['main']][state['sub']] = []
                
            # 避免重复添加
            existing = next((c for c in result[state['main']][state['sub']] 
                          if c["name"] == company_data["name"]), None)
            if not existing:
                result[state['main']][state['sub']].append(company_data)
    
    return result

def save_to_js(data, output_file):
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(f"const stockData = {json.dumps(data, ensure_ascii=False, indent=2)};\n")
        f.write("export default stockData;")

if __name__ == "__main__":
    input_file = "stock-list.xlsx"
    output_file = "stockData.js"
    
    # 初始化日志文件
    log_file = open("log.txt", "w", encoding="utf-8")
    
    print("开始处理股票数据...")
    structured_data = read_excel_to_dict(input_file)
    
    print("正在保存结果...")
    save_to_js(structured_data, output_file)
    
    print(f"数据已成功导出到 {output_file}")
    
    # 关闭日志文件
    if log_file:
        log_file.close()
